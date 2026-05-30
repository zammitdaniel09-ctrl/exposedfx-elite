import re
import sqlite3
import time
import asyncio
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DEST_CHAT = -1003952162034
DEST_TOPIC = 5
TZ_NAME = "Europe/Malta"

BLACK = "0B0B0B"
DARK = "15120A"
GOLD = "D9A441"
GOLD2 = "F4C542"
WHITE = "F2F2F2"
GREEN = "37B24D"
RED = "FF3B30"


def parse_stats_result(text):
    text = (text or "").replace(",", " ")
    upper = re.sub(r"\s+", " ", text.upper()).strip()
    if not upper:
        return None
    if any(x in upper for x in ["PIP VALUE", "WHAT IS A PIP", "EXAMPLE"]):
        return None
    pip_vals = [float(m.group(1)) for m in re.finditer(r"([+-]?\d+(?:\.\d+)?)\s*(?:PIP|PIPS)\b", upper)]
    pips = max(pip_vals, key=lambda x: abs(x)) if pip_vals else None
    if any(x in upper for x in ["SL HIT", "STOP LOSS HIT", "HIT SL", "STOPPED OUT"]):
        return {"status": "LOSS", "pips": -abs(pips) if pips is not None else 0.0}
    if pips is None:
        return None
    if pips < 0:
        return {"status": "LOSS", "pips": pips}
    if any(x in upper for x in ["PIPS", "TP HIT", "PROFIT", "CLOSED", "SECURED", "BANKED", "BOOKED", "CAUGHT", "SMASHED"]):
        return {"status": "WIN", "pips": abs(pips)}
    return None


class WeeklyStats:
    def __init__(self, data_dir):
        self.data_dir = Path(data_dir)
        self.db_path = self.data_dir / "stats.sqlite"
        self.report_dir = self.data_dir / "reports"
        self.report_dir.mkdir(parents=True, exist_ok=True)
        self.tz = ZoneInfo(TZ_NAME)
        self.init_db()

    def connect(self):
        con = sqlite3.connect(str(self.db_path))
        con.row_factory = sqlite3.Row
        return con

    def init_db(self):
        con = self.connect()
        con.execute("CREATE TABLE IF NOT EXISTS results(id INTEGER PRIMARY KEY AUTOINCREMENT, msg_key TEXT UNIQUE, source TEXT, status TEXT, pips REAL, raw_text TEXT, created_at REAL)")
        con.execute("CREATE TABLE IF NOT EXISTS reports(id INTEGER PRIMARY KEY AUTOINCREMENT, week_no INTEGER, start_ts REAL, end_ts REAL, total_pips REAL, file_path TEXT, sent_at REAL)")
        con.execute("CREATE TABLE IF NOT EXISTS meta(key TEXT PRIMARY KEY, value TEXT)")
        con.commit(); con.close()

    def log_message(self, route, message, text):
        result = parse_stats_result(text)
        if not result:
            return None
        key = f"{route.get('source_chat')}:{getattr(message, 'id', '')}"
        con = self.connect()
        try:
            con.execute("INSERT INTO results(msg_key, source, status, pips, raw_text, created_at) VALUES(?,?,?,?,?,?)", (key, route.get("name", "Unknown"), result["status"], result["pips"], text, time.time()))
            con.commit(); return result
        except sqlite3.IntegrityError:
            return None
        finally:
            con.close()

    def period(self):
        now = datetime.now(self.tz)
        today = now.replace(hour=0, minute=0, second=0, microsecond=0)
        sunday = today - timedelta(days=(today.weekday() - 6) % 7)
        start = sunday - timedelta(days=7)
        end = sunday
        return start.timestamp(), end.timestamp(), start.strftime("%d %b %Y"), (end - timedelta(seconds=1)).strftime("%d %b %Y")

    def should_send(self):
        now = datetime.now(self.tz)
        if now.weekday() != 6 or now.hour != 0:
            return False
        start_ts, _, _, _ = self.period()
        con = self.connect(); row = con.execute("SELECT value FROM meta WHERE key=?", (f"sent:{int(start_ts)}",)).fetchone(); con.close()
        return row is None

    def mark_sent(self, start_ts):
        con = self.connect(); con.execute("INSERT OR REPLACE INTO meta(key,value) VALUES(?,?)", (f"sent:{int(start_ts)}", str(time.time()))); con.commit(); con.close()

    def next_week_no(self):
        con = self.connect(); row = con.execute("SELECT MAX(week_no) AS n FROM reports").fetchone(); con.close()
        return int(row["n"] or 0) + 1

    def style_sheet(self, ws):
        ws.sheet_view.showGridLines = False
        thin = Side(style="thin", color=GOLD)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.font = Font(name="Aptos", color=WHITE)
                cell.fill = PatternFill("solid", fgColor=BLACK)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def build_report(self):
        start_ts, end_ts, start_label, end_label = self.period()
        con = self.connect()
        rows = con.execute("SELECT * FROM results WHERE created_at>=? AND created_at<? ORDER BY created_at", (start_ts, end_ts)).fetchall()
        week_no = self.next_week_no()
        wins = sum(1 for r in rows if r["status"] == "WIN"); losses = sum(1 for r in rows if r["status"] == "LOSS")
        total = wins + losses; total_pips = sum(float(r["pips"] or 0) for r in rows)
        avg = total_pips / total if total else 0; wr = wins / total * 100 if total else 0
        by_source = {}
        for r in rows:
            by_source[r["source"]] = by_source.get(r["source"], 0) + float(r["pips"] or 0)
        best_source = max(by_source.items(), key=lambda x: x[1])[0] if by_source else "N/A"
        best_trade = max([r for r in rows if r["status"] == "WIN"], key=lambda r: float(r["pips"] or 0), default=None)
        best_trade_text = f"{best_trade['source']} | {float(best_trade['pips']):.1f} pips" if best_trade else "N/A"

        wb = Workbook(); ws = wb.active; ws.title = "Summary"; tr = wb.create_sheet("Trade Log")
        for s in [ws, tr]:
            s.sheet_view.showGridLines = False
        ws.merge_cells("A1:H2"); ws["A1"] = f"ExposedFX Preview Weekly Stats - Week {week_no}"
        ws["A1"].font = Font(name="Aptos Display", size=24, bold=True, color=GOLD2); ws["A1"].fill = PatternFill("solid", fgColor=BLACK); ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A3:H3"); ws["A3"] = f"Period: {start_label} - {end_label}"; ws["A3"].font = Font(size=13, color=WHITE); ws["A3"].fill = PatternFill("solid", fgColor=DARK); ws["A3"].alignment = Alignment(horizontal="center")
        cards = [("WINS", wins), ("LOSSES", losses), ("WIN RATE", f"{wr:.2f}%"), ("TOTAL PIPS", f"{total_pips:.1f}"), ("AVERAGE PIPS", f"{avg:.1f}"), ("BEST GROUP", best_source), ("BEST TRADE", best_trade_text), ("STATUS", "CONFIDENTIAL")]
        pos = ["A5:B7", "C5:D7", "E5:F7", "G5:H7", "A9:B11", "C9:D11", "E9:F11", "G9:H11"]
        for area, (title, val) in zip(pos, cards):
            ws.merge_cells(area); c = ws[area.split(":")[0]]; c.value = f"{title}\n{val}"; c.fill = PatternFill("solid", fgColor=BLACK); c.font = Font(size=15, bold=True, color=GOLD2); c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        headers = ["Date", "Source", "Result", "Pips", "Message"]
        for i, h in enumerate(headers, 1):
            cell = tr.cell(1, i, h); cell.fill = PatternFill("solid", fgColor=GOLD); cell.font = Font(bold=True, color=BLACK); cell.alignment = Alignment(horizontal="center")
        for r_i, r in enumerate(rows, 2):
            dt = datetime.fromtimestamp(float(r["created_at"]), self.tz).strftime("%d %b %Y %H:%M")
            vals = [dt, r["source"], r["status"], float(r["pips"] or 0), r["raw_text"]]
            for c_i, val in enumerate(vals, 1):
                cell = tr.cell(r_i, c_i, val); cell.fill = PatternFill("solid", fgColor=BLACK); cell.font = Font(color=WHITE); cell.alignment = Alignment(horizontal="center", wrap_text=True)
            tr.cell(r_i, 3).font = Font(bold=True, color=GREEN if r["status"] == "WIN" else RED)
            tr.cell(r_i, 4).font = Font(bold=True, color=GREEN if float(r["pips"] or 0) >= 0 else RED)
        for s in [ws, tr]:
            for col in range(1, 9):
                s.column_dimensions[get_column_letter(col)].width = 18
            for row in range(1, s.max_row + 1):
                s.row_dimensions[row].height = 28
        tr.column_dimensions[5].width = 55; tr.freeze_panes = "A2"
        file_path = self.report_dir / f"ExposedFX_Preview_Week_{week_no}.xlsx"; wb.save(file_path)
        con.execute("INSERT INTO reports(week_no,start_ts,end_ts,total_pips,file_path,sent_at) VALUES(?,?,?,?,?,?)", (week_no, start_ts, end_ts, total_pips, str(file_path), time.time()))
        con.commit(); best = con.execute("SELECT week_no,total_pips FROM reports ORDER BY total_pips DESC LIMIT 1").fetchone(); con.close()
        caption = f"ExposedFX Preview Weekly Stats - Week {week_no}\nPeriod: {start_label} - {end_label}\n\nWins: {wins}\nLosses: {losses}\nWin Rate: {wr:.2f}%\nTotal Pips: {total_pips:.1f}\nAverage Pips: {avg:.1f}\nBest Group: {best_source}\nBest Trade: {best_trade_text}\nBest Week: Week {best['week_no']} | {float(best['total_pips']):.1f} pips\n\nPremium spreadsheet attached."
        return start_ts, file_path, caption

    async def loop(self, client):
        while True:
            try:
                if self.should_send():
                    start_ts, file_path, caption = self.build_report()
                    await client.send_file(DEST_CHAT, file_path, caption=caption, reply_to=DEST_TOPIC)
                    self.mark_sent(start_ts)
            except Exception as exc:
                print(f"[stats reporter failed] {exc}")
            await asyncio.sleep(60)
